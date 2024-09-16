from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
from libs.docx_manipulator import DocxGenerator
from utils import *
from paralelismo import *
from validacao import validar_entrada
from io import BytesIO
from flask import send_file
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side

def gerar_planilha_excel(dados_equacao):
    wb = Workbook()

    def ajustar_largura_colunas(ws):
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)  # Obtém a letra da coluna
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

    def centralizar_conteudo(ws):
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    def adicionar_bordas(ws):
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border

    # Abas da planilha
    if dados_equacao['opcao'] == 'valor 1':
        ws1 = wb.active
        ws1.title = "Concentração sem Correção"
        ws1.append(['N°', 'Massa padrão (mg)', 'Fator de equivalência (g/mol)', 'Volume de solução (mol/L)', 'Concentração Final'])
        
        for i, resultado in enumerate(dados_equacao['variaveis']):
            ws1.append([i + 1] + resultado + [dados_equacao['resultados'][i]])

        ajustar_largura_colunas(ws1)
        centralizar_conteudo(ws1)
        adicionar_bordas(ws1)

    elif dados_equacao['opcao'] == 'valor 2':
        ws2 = wb.active
        ws2.title = "Concentração com Correção"
        ws2.append(['N°', 'Massa padrão (mg)', 'Fator de correção (g/mol)', 'Fator de equivalência (g/mol)', 'Volume NAOH (mol/L)', 'Concentração Final'])
        
        for i, resultado in enumerate(dados_equacao['variaveis']):
            ws2.append([i + 1] + resultado + [dados_equacao['resultados'][i]])

        ajustar_largura_colunas(ws2)
        centralizar_conteudo(ws2)
        adicionar_bordas(ws2)

    elif dados_equacao['opcao'] == 'valor 3':
        ws3 = wb.active
        ws3.title = "Concentração Molar"
        ws3.append(['N°', 'Massa soluto (mg)', 'Massa molar (g/mol)', 'Volume solução (mol/L)', 'Concentração Final'])
        
        for i, resultado in enumerate(dados_equacao['variaveis']):
            ws3.append([i + 1] + resultado + [dados_equacao['resultados'][i]])

        ajustar_largura_colunas(ws3)
        centralizar_conteudo(ws3)
        adicionar_bordas(ws3)

    # Salva o arquivo
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


app = Flask(__name__)


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == 'GET':
        return render_template('index.html')
    
    if request.method == "POST":
        selected_eq = request.form.get('eq_select').strip()
        quantity_eq = request.form.get('eq_quantity').strip()
        return redirect(url_for("equacoes", selected_eq=selected_eq, quantity_eq=quantity_eq))

@app.route("/equacoes", methods=["GET","POST"])
def equacoes():
    global selected_eq_atual
    global quantity_eq_atual
    global valores_totais

    if request.method == "GET":
        selected_eq = request.args.get("selected_eq")
        quantity_eq = request.args.get("quantity_eq")
        selected_eq_atual = selected_eq
        quantity_eq_atual = quantity_eq
        base = select_base(selected_eq_atual, quantity_eq_atual)
        return render_template('equacoes.html', result=base, error_msg=None)
    
    if request.method == "POST":
        valores_totais = []
        error_msg = None
        
        if selected_eq_atual == "valor 3":
            for i in range(int(quantity_eq_atual)):
                i+=1
                ms_result = request.form.get(f"MS_result_{i}")
                mm_result = request.form.get(f"MM_result_{i}")
                vs_result = request.form.get(f"VS_result_{i}")
                
                valid_ms, msg_ms = validar_entrada(ms_result)
                valid_mm, msg_mm = validar_entrada(mm_result)
                valid_vs, msg_vs = validar_entrada(vs_result)
                
                if not (valid_ms and valid_mm and valid_vs):
                    error_msg = f"Erros encontrados: {msg_ms} {msg_mm} {msg_vs}"
                    base = select_base(selected_eq_atual, quantity_eq_atual)
                    return render_template('equacoes.html', result=base, error_msg=error_msg)
                
                valores_totais.append([float(ms_result), float(mm_result), float(vs_result)])
            
            return redirect("/resultados")
        
        else:
            for i in range(int(quantity_eq_atual)):
                i+=1
                mp_result = request.form.get(f"MP_result_{i}")
                fe_result = request.form.get(f"FE_result_{i}")
                vn_result = request.form.get(f"VN_result_{i}")
                
                valid_mp, msg_mp = validar_entrada(mp_result)
                valid_fe, msg_fe = validar_entrada(fe_result)
                valid_vn, msg_vn = validar_entrada(vn_result)
                
                if not (valid_mp and valid_fe and valid_vn):
                    error_msg = f"Erros encontrados: {msg_mp} {msg_fe} {msg_vn}"
                    if selected_eq_atual == "valor 2":
                        fc_result = request.form.get(f"FC_result_{i}")
                        valid_fc, msg_fc = validar_entrada(fc_result)
                        error_msg += f" {msg_fc}"
                    
                    base = select_base(selected_eq_atual, quantity_eq_atual)
                    return render_template('equacoes.html', result=base, error_msg=error_msg)
                
                if selected_eq_atual == "valor 2":
                    fc_result = request.form.get(f"FC_result_{i}")
                    valores_totais.append([float(mp_result), float(fe_result), float(vn_result), float(fc_result)])
                else:
                    valores_totais.append([float(mp_result), float(fe_result), float(vn_result)])
            
            return redirect("/resultados")

@app.route("/resultados", methods=["GET"])
def resultados():
    selected_eq = selected_eq_atual
    variaveis = valores_totais
    
    if selected_eq == "valor 1":
        resultado = thread_concentracao_sem_correcao(variaveis)
    elif selected_eq == "valor 2":
        resultado = thread_concentracao_com_correcao(variaveis)
    elif selected_eq == "valor 3":
        resultado = thread_concentracao_molar(variaveis)
    
    dados_equacao = {
        'opcao': selected_eq,
        'variaveis': variaveis,
        'resultados': resultado,
    }

    return render_template('resultados.html', dados_equacao=dados_equacao)

@app.route("/gerar-docx", methods=['POST'])
def relatorio_docx():
    dados_equacao = request.get_json()
    document = DocxGenerator()

    records = dados_equacao['variaveis']

    for indice, variaveis in enumerate(records):
        variaveis.insert(0, indice+1)
        variaveis.append(dados_equacao['resultados'][indice])

    if dados_equacao['opcao'] == 'valor 1':
        document.add_table_concentracao_sem_correcao(records)
    elif dados_equacao['opcao'] == 'valor 2':
        document.add_table_concentracao_com_correcao(records)
    elif dados_equacao['opcao'] == 'valor 3':
        document.add_table_concentracao_molar(records)
    
    filename = 'exemplo'
    document.save_doc(filename)

    return jsonify({'redirect': url_for('dowload_docx', filename=filename)})

@app.route('/dowload-docx/<filename>')
def dowload_docx(filename):
    return redirect(url_for('static', filename=f'docs/{filename}.docx'))

@app.route("/gerar-planilha", methods=['POST'])
def gerar_planilha():
    dados_equacao = request.get_json()
    buffer = gerar_planilha_excel(dados_equacao)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name='resultados_calculos.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


if __name__ == '__main__':
    app.run(debug=True)
