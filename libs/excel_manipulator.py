# C:\Users\cleys\OneDrive\Área de Trabalho\Projeto-4\libs\excel_manipulator.py

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side
from io import BytesIO

def gerar_planilha_excel(dados_equacao):
    wb = Workbook()

    def ajustar_largura_colunas(ws):
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)  # Obtém a letra da coluna
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

    def centralizar_conteudo(ws):
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    def adicionar_bordas(ws):
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border

    # Abas da planilha
    if dados_equacao['opcao'] == 'valor 1':
        ws1 = wb.active
        ws1.title = "Concentração sem Correção"
        ws1.append(['N°', 'Massa padrão (mg)', 'Fator de equivalência (g/mol)', 'Volume de solução (mol/L)', 'Concentração Final'])
        
        for i, resultado in enumerate(dados_equacao['variaveis']):
            ws1.append([i + 1] + resultado + [dados_equacao['resultados'][i]])

        ajustar_largura_colunas(ws1)
        centralizar_conteudo(ws1)
        adicionar_bordas(ws1)

    elif dados_equacao['opcao'] == 'valor 2':
        ws2 = wb.active
        ws2.title = "Concentração com Correção"
        ws2.append(['N°', 'Massa padrão (mg)', 'Fator de correção (g/mol)', 'Fator de equivalência (g/mol)', 'Volume NAOH (mol/L)', 'Concentração Final'])
        
        for i, resultado in enumerate(dados_equacao['variaveis']):
            ws2.append([i + 1] + resultado + [dados_equacao['resultados'][i]])

        ajustar_largura_colunas(ws2)
        centralizar_conteudo(ws2)
        adicionar_bordas(ws2)

    elif dados_equacao['opcao'] == 'valor 3':
        ws3 = wb.active
        ws3.title = "Concentração Molar"
        ws3.append(['N°', 'Massa soluto (mg)', 'Massa molar (g/mol)', 'Volume solução (mol/L)', 'Concentração Final'])
        
        for i, resultado in enumerate(dados_equacao['variaveis']):
            ws3.append([i + 1] + resultado + [dados_equacao['resultados'][i]])

        ajustar_largura_colunas(ws3)
        centralizar_conteudo(ws3)
        adicionar_bordas(ws3)

    # Salva o arquivo
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
